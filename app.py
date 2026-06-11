from datetime import datetime, timedelta
import json, os, re

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None
    load_workbook = None

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
USERS_DIR = os.path.join(BASE_DIR, 'usuarios')
DATA_DIR = os.path.join(BASE_DIR, 'dados')
MATCHES_FILE = os.path.join(DATA_DIR, 'jogos.json')
PORT = int(os.environ.get('PORT', 8080))
ADMIN_USER = 'admin'
ADMIN_PASS = 'admin2026'

GROUPS = {
 'A':['México','África do Sul','Coreia do Sul','Tchéquia'],
 'B':['Bósnia e Herzegovina','Canadá','Qatar','Suíça'],
 'C':['Brasil','Marrocos','Escócia','Haiti'],
 'D':['Estados Unidos','Turquia','Austrália','Paraguai'],
 'E':['Alemanha','Equador','Curaçao','Costa do Marfim'],
 'F':['Países Baixos','Japão','Suécia','Tunísia'],
 'G':['Bélgica','Irã','Egito','Nova Zelândia'],
 'H':['Espanha','Uruguai','Arábia Saudita','Cabo Verde'],
 'I':['França','Senegal','Noruega','Iraque'],
 'J':['Argentina','Áustria','Argélia','Jordânia'],
 'K':['Portugal','Colômbia','Uzbequistão','República Democrática do Congo'],
 'L':['Inglaterra','Croácia','Panamá','Gana']
}

def now_br():
    return datetime.now().strftime('%d/%m/%Y %H:%M:%S')

def safe_login(login):
    login = (login or '').strip().lower()
    return re.sub(r'[^a-z0-9._-]', '_', login)

def user_path(login):
    return os.path.join(USERS_DIR, f'{safe_login(login)}.xlsx')

def ensure_dirs():
    os.makedirs(USERS_DIR, exist_ok=True)
    os.makedirs(DATA_DIR, exist_ok=True)

SCHEDULE_DATA = [
    # Horários em Brasília/São Paulo (BRT, UTC-3). Base pesquisada em FIFA/FOX/Sky Sports e mantida editável pelo admin.
    (1,'A','México','África do Sul','2026-06-11T16:00','Mexico City Stadium','Mexico City'),
    (2,'A','Coreia do Sul','Tchéquia','2026-06-11T23:00','Guadalajara Stadium','Guadalajara'),
    (3,'B','Canadá','Bósnia e Herzegovina','2026-06-12T16:00','Toronto Stadium','Toronto'),
    (4,'D','Estados Unidos','Paraguai','2026-06-12T22:00','Los Angeles Stadium','Los Angeles'),
    (5,'B','Qatar','Suíça','2026-06-13T16:00','San Francisco Bay Area Stadium','San Francisco Bay Area'),
    (6,'C','Brasil','Marrocos','2026-06-13T19:00','New York New Jersey Stadium','New York/New Jersey'),
    (7,'C','Haiti','Escócia','2026-06-13T22:00','Boston Stadium','Boston'),
    (8,'D','Austrália','Turquia','2026-06-14T01:00','BC Place Vancouver','Vancouver'),
    (9,'E','Alemanha','Curaçao','2026-06-14T14:00','Houston Stadium','Houston'),
    (10,'F','Países Baixos','Japão','2026-06-14T17:00','Dallas Stadium','Dallas'),
    (11,'E','Costa do Marfim','Equador','2026-06-14T20:00','Philadelphia Stadium','Philadelphia'),
    (12,'F','Suécia','Tunísia','2026-06-14T23:00','Monterrey Stadium','Monterrey'),
    (13,'H','Espanha','Cabo Verde','2026-06-15T13:00','Atlanta Stadium','Atlanta'),
    (14,'G','Bélgica','Egito','2026-06-15T16:00','Seattle Stadium','Seattle'),
    (15,'H','Arábia Saudita','Uruguai','2026-06-15T19:00','Miami Stadium','Miami'),
    (16,'G','Irã','Nova Zelândia','2026-06-15T22:00','Los Angeles Stadium','Los Angeles'),
    (17,'I','França','Senegal','2026-06-16T16:00','New York New Jersey Stadium','New York/New Jersey'),
    (18,'I','Iraque','Noruega','2026-06-16T19:00','Boston Stadium','Boston'),
    (19,'J','Argentina','Argélia','2026-06-16T22:00','Kansas City Stadium','Kansas City'),
    (20,'J','Áustria','Jordânia','2026-06-17T01:00','San Francisco Bay Area Stadium','San Francisco Bay Area'),
    (21,'K','Portugal','República Democrática do Congo','2026-06-17T14:00','Houston Stadium','Houston'),
    (22,'L','Inglaterra','Croácia','2026-06-17T17:00','Dallas Stadium','Dallas'),
    (23,'L','Gana','Panamá','2026-06-17T20:00','Toronto Stadium','Toronto'),
    (24,'K','Uzbequistão','Colômbia','2026-06-17T23:00','Mexico City Stadium','Mexico City'),
    (25,'A','Tchéquia','África do Sul','2026-06-18T13:00','Atlanta Stadium','Atlanta'),
    (26,'B','Suíça','Bósnia e Herzegovina','2026-06-18T16:00','Los Angeles Stadium','Los Angeles'),
    (27,'A','México','Coreia do Sul','2026-06-18T19:00','Guadalajara Stadium','Guadalajara'),
    (28,'B','Canadá','Qatar','2026-06-18T22:00','Kansas City Stadium','Kansas City'),
    (29,'C','Brasil','Haiti','2026-06-19T21:30','Philadelphia Stadium','Philadelphia'),
    (30,'D','Turquia','Paraguai','2026-06-20T00:00','San Francisco Bay Area Stadium','San Francisco Bay Area'),
    (31,'F','Países Baixos','Suécia','2026-06-20T14:00','Houston Stadium','Houston'),
    (32,'E','Alemanha','Costa do Marfim','2026-06-20T17:00','Toronto Stadium','Toronto'),
    (33,'E','Equador','Curaçao','2026-06-20T21:00','Kansas City Stadium','Kansas City'),
    (34,'F','Tunísia','Japão','2026-06-21T01:00','Monterrey Stadium','Monterrey'),
    (35,'H','Espanha','Arábia Saudita','2026-06-21T13:00','Atlanta Stadium','Atlanta'),
    (36,'G','Bélgica','Irã','2026-06-21T16:00','Los Angeles Stadium','Los Angeles'),
    (37,'H','Uruguai','Cabo Verde','2026-06-21T19:00','Miami Stadium','Miami'),
    (38,'G','Nova Zelândia','Egito','2026-06-21T22:00','BC Place Vancouver','Vancouver'),
    (39,'J','Argentina','Áustria','2026-06-22T14:00','Dallas Stadium','Dallas'),
    (40,'I','França','Iraque','2026-06-22T18:00','Philadelphia Stadium','Philadelphia'),
    (41,'I','Noruega','Senegal','2026-06-22T21:00','Toronto Stadium','Toronto'),
    (42,'J','Jordânia','Argélia','2026-06-23T00:00','San Francisco Bay Area Stadium','San Francisco Bay Area'),
    (43,'K','Portugal','Uzbequistão','2026-06-23T14:00','Houston Stadium','Houston'),
    (44,'L','Inglaterra','Gana','2026-06-23T17:00','Boston Stadium','Boston'),
    (45,'L','Panamá','Croácia','2026-06-23T20:00','Boston Stadium','Boston'),
    (46,'K','Colômbia','República Democrática do Congo','2026-06-23T23:00','Guadalajara Stadium','Guadalajara'),
    (47,'B','Suíça','Canadá','2026-06-24T16:00','BC Place Vancouver','Vancouver'),
    (48,'B','Bósnia e Herzegovina','Qatar','2026-06-24T16:00','Seattle Stadium','Seattle'),
    (49,'C','Marrocos','Haiti','2026-06-24T19:00','Atlanta Stadium','Atlanta'),
    (50,'C','Escócia','Brasil','2026-06-24T19:00','Miami Stadium','Miami'),
    (51,'A','África do Sul','Coreia do Sul','2026-06-24T22:00','Monterrey Stadium','Monterrey'),
    (52,'A','Tchéquia','México','2026-06-24T22:00','Mexico City Stadium','Mexico City'),
    (53,'E','Curaçao','Costa do Marfim','2026-06-25T17:00','Philadelphia Stadium','Philadelphia'),
    (54,'E','Equador','Alemanha','2026-06-25T17:00','New York New Jersey Stadium','New York/New Jersey'),
    (55,'F','Tunísia','Países Baixos','2026-06-25T20:00','Kansas City Stadium','Kansas City'),
    (56,'F','Japão','Suécia','2026-06-25T20:00','Dallas Stadium','Dallas'),
    (57,'D','Turquia','Estados Unidos','2026-06-25T23:00','Los Angeles Stadium','Los Angeles'),
    (58,'D','Paraguai','Austrália','2026-06-25T23:00','San Francisco Bay Area Stadium','San Francisco Bay Area'),
    (59,'I','Noruega','França','2026-06-26T16:00','Boston Stadium','Boston'),
    (60,'I','Senegal','Iraque','2026-06-26T16:00','Toronto Stadium','Toronto'),
    (61,'H','Cabo Verde','Arábia Saudita','2026-06-26T21:00','Houston Stadium','Houston'),
    (62,'H','Uruguai','Espanha','2026-06-26T21:00','Guadalajara Stadium','Guadalajara'),
    (63,'G','Nova Zelândia','Bélgica','2026-06-27T00:00','BC Place Vancouver','Vancouver'),
    (64,'G','Egito','Irã','2026-06-27T00:00','Seattle Stadium','Seattle'),
    (65,'L','Panamá','Inglaterra','2026-06-27T18:00','New York New Jersey Stadium','New York/New Jersey'),
    (66,'L','Croácia','Gana','2026-06-27T18:00','Philadelphia Stadium','Philadelphia'),
    (67,'K','Colômbia','Portugal','2026-06-27T20:30','Miami Stadium','Miami'),
    (68,'K','República Democrática do Congo','Uzbequistão','2026-06-27T20:30','Atlanta Stadium','Atlanta'),
    (69,'J','Argélia','Áustria','2026-06-27T23:00','Kansas City Stadium','Kansas City'),
    (70,'J','Jordânia','Argentina','2026-06-27T23:00','Dallas Stadium','Dallas'),
    (71,'D','Estados Unidos','Austrália','2026-06-19T22:00','Seattle Stadium','Seattle'),
    (72,'C','Escócia','Marrocos','2026-06-19T19:00','Boston Stadium','Boston'),
]

def build_default_matches():
    arr=[]
    for mid, g, home, away, kickoff, stadium, city in sorted(SCHEDULE_DATA, key=lambda x: x[0]):
        arr.append({
            'id': mid, 'group': g, 'round': 1,
            'home': home, 'away': away,
            'kickoff': kickoff,
            'stadium': stadium,
            'city': city,
            'hg': None, 'ag': None
        })
    return arr

def load_matches():
    ensure_dirs()
    defaults = {int(m['id']): m for m in build_default_matches()}
    if not os.path.exists(MATCHES_FILE):
        matches = list(defaults.values())
        with open(MATCHES_FILE, 'w', encoding='utf-8') as f: json.dump(matches, f, ensure_ascii=False, indent=2)
        return matches
    with open(MATCHES_FILE, 'r', encoding='utf-8') as f:
        matches = json.load(f)
    changed = False
    for m in matches:
        d = defaults.get(int(m.get('id', 0)))
        if d:
            for key in ['stadium','city']:
                if not m.get(key):
                    m[key] = d.get(key, '')
                    changed = True
    if changed:
        save_matches(matches)
    return matches

def save_matches(matches):
    ensure_dirs()
    with open(MATCHES_FILE, 'w', encoding='utf-8') as f: json.dump(matches, f, ensure_ascii=False, indent=2)

def style_sheet(ws):
    header_fill = PatternFill('solid', fgColor='17365D')
    header_font = Font(color='FFFFFF', bold=True)
    thin = Side(style='thin', color='D9E2F3')
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(left=thin,right=thin,top=thin,bottom=thin)
            cell.alignment = Alignment(vertical='center')
    for cell in ws[1]:
        cell.fill = header_fill; cell.font = header_font; cell.alignment = Alignment(horizontal='center')
    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len+3, 45)

def create_user_file(login, password, nome=''):
    if Workbook is None: raise RuntimeError('openpyxl não instalado')
    path = user_path(login)
    if os.path.exists(path): return False, 'Usuário já existe.'
    wb = Workbook()
    ws = wb.active; ws.title = 'USUARIO'
    ws.append(['Login','Senha','Nome','Criado em'])
    ws.append([safe_login(login), password, nome, now_br()])
    log = wb.create_sheet('LOG_APOSTAS')
    log.append(['Data/Hora','Login','Ação','Jogo ID','Grupo','Time Casa','Time Fora','Início do Jogo','Estádio','Cidade','Aposta Casa Anterior','Aposta Fora Anterior','Aposta Casa Nova','Aposta Fora Nova','Status','Observação'])
    bets = wb.create_sheet('APOSTAS_ATUAIS')
    bets.append(['Jogo ID','Grupo','Time Casa','Time Fora','Início do Jogo','Estádio','Cidade','Aposta Casa Atual','Aposta Fora Atual','Última alteração'])
    for s in wb.worksheets: style_sheet(s)
    wb.save(path)
    return True, path

def read_user_password(login):
    path = user_path(login)
    if not os.path.exists(path): return None, 'Usuário não encontrado. A planilha dele não existe na pasta usuarios.'
    wb = load_workbook(path)
    ws = wb['USUARIO']
    return str(ws['B2'].value or ''), None

def validate_user(login, password):
    saved, err = read_user_password(login)
    if err: return False, err
    if str(password) != saved: return False, 'Senha incorreta para este usuário.'
    return True, 'Login validado com sucesso.'

def is_locked(match):
    kickoff = datetime.strptime(match['kickoff'], '%Y-%m-%dT%H:%M')
    return datetime.now() >= kickoff - timedelta(minutes=5)

def update_bet(login, match_id, bet_home, bet_away):
    matches = load_matches()
    match = next((m for m in matches if int(m['id']) == int(match_id)), None)
    if not match: return False, 'Jogo não encontrado.'
    if is_locked(match): return False, 'Apostas bloqueadas: faltam menos de 5 minutos para o início ou o jogo já começou.'

    bet_home = int(bet_home)
    bet_away = int(bet_away)
    path = user_path(login)
    wb = load_workbook(path)
    log = wb['LOG_APOSTAS']
    bets = wb['APOSTAS_ATUAIS']

    action = 'CRIACAO_APOSTA'
    row_found = None
    old_home = ''
    old_away = ''

    for r in range(2, bets.max_row + 1):
        current_id = bets.cell(r, 1).value
        if current_id is not None and int(current_id) == int(match_id):
            row_found = r
            break

    # A aba APOSTAS_ATUAIS mantém somente a aposta válida atual do jogo.
    # A aba LOG_APOSTAS nunca é sobrescrita: cada salvamento vira uma nova linha de histórico.
    if row_found:
        action = 'ALTERACAO_APOSTA'
        old_home = bets.cell(row_found, 8).value
        old_away = bets.cell(row_found, 9).value
        bets.cell(row_found, 8).value = bet_home
        bets.cell(row_found, 9).value = bet_away
        bets.cell(row_found, 10).value = now_br()
    else:
        bets.append([
            match['id'], match['group'], match['home'], match['away'], match['kickoff'],
            match.get('stadium', ''), match.get('city', ''), bet_home, bet_away, now_br()
        ])

    log.append([
        now_br(), safe_login(login), action, match['id'], match['group'], match['home'], match['away'],
        match['kickoff'], match.get('stadium', ''), match.get('city', ''), old_home, old_away,
        bet_home, bet_away, 'OK',
        'Primeira aposta gravada' if action == 'CRIACAO_APOSTA' else 'Alteração registrada sem apagar o histórico anterior'
    ])

    for s in [log, bets]:
        style_sheet(s)
    wb.save(path)
    return True, f'Aposta salva na planilha usuarios/{safe_login(login)}.xlsx. Histórico adicionado em LOG_APOSTAS.'

def load_user_bets(login):
    path = user_path(login)
    if not os.path.exists(path): return []
    wb = load_workbook(path, data_only=True)
    ws = wb['APOSTAS_ATUAIS']
    bets=[]
    for r in range(2, ws.max_row+1):
        if ws.cell(r,1).value is not None:
            bets.append({'match_id': int(ws.cell(r,1).value), 'home': ws.cell(r,8).value, 'away': ws.cell(r,9).value, 'updated': ws.cell(r,10).value})
    return bets

def calc_standings(matches):
    table = {g:{t:{'Grupo':g,'Seleção':t,'Pts':0,'J':0,'V':0,'E':0,'D':0,'GP':0,'GC':0,'SG':0} for t in teams} for g,teams in GROUPS.items()}
    for m in matches:
        if m.get('hg') in [None,''] or m.get('ag') in [None,'']: continue
        hg, ag = int(m['hg']), int(m['ag']); g=m['group']; home=m['home']; away=m['away']
        h=table[g][home]; a=table[g][away]
        h['J']+=1; a['J']+=1; h['GP']+=hg; h['GC']+=ag; a['GP']+=ag; a['GC']+=hg
        if hg>ag: h['V']+=1; a['D']+=1; h['Pts']+=3
        elif hg<ag: a['V']+=1; h['D']+=1; a['Pts']+=3
        else: h['E']+=1; a['E']+=1; h['Pts']+=1; a['Pts']+=1
        h['SG']=h['GP']-h['GC']; a['SG']=a['GP']-a['GC']
    return {g: sorted(list(v.values()), key=lambda x:(-x['Pts'],-x['SG'],-x['GP'],x['GC'],x['Seleção'])) for g,v in table.items()}


from flask import Flask, request, jsonify, send_from_directory

# WSGI app exigido pelo Render/Gunicorn:
# Start Command no Render: gunicorn app:app
app = Flask(__name__, static_folder=BASE_DIR, static_url_path='')

@app.before_request
def _prepare_app():
    ensure_dirs()
    load_matches()

@app.route('/')
def index():
    index_path = os.path.join(BASE_DIR, 'index.html')
    if os.path.exists(index_path):
        return send_from_directory(BASE_DIR, 'index.html')
    return '<h1>Bolão Copa 2026 Pluma</h1><p>Arquivo index.html não encontrado na raiz do projeto.</p>', 200

@app.route('/<path:filename>')
def static_files(filename):
    # Serve arquivos do site: style.css, script.js, imagens, etc.
    # Bloqueia acesso direto às planilhas e pastas internas.
    blocked_prefixes = ('usuarios/', 'dados/', '.git/', '__pycache__/')
    normalized = filename.replace('\\', '/')
    if normalized.startswith(blocked_prefixes) or normalized.endswith('.xlsx'):
        return jsonify({'ok': False, 'erro': 'Acesso não permitido.'}), 403
    target = os.path.join(BASE_DIR, filename)
    if os.path.exists(target) and os.path.isfile(target):
        return send_from_directory(BASE_DIR, filename)
    return jsonify({'ok': False, 'erro': 'Arquivo não encontrado.'}), 404

@app.route('/api/jogos', methods=['GET'])
def api_jogos():
    matches = load_matches()
    return jsonify({'ok': True, 'matches': matches, 'standings': calc_standings(matches)})

@app.route('/api/apostas/<login>', methods=['GET'])
def api_apostas(login):
    return jsonify({'ok': True, 'bets': load_user_bets(login)})

@app.route('/api/criar-usuario', methods=['POST'])
def api_criar_usuario():
    data = request.get_json(silent=True) or {}
    login = safe_login(data.get('login'))
    password = str(data.get('password', '')).strip()
    nome = str(data.get('nome', '')).strip()

    if len(login) < 3:
        return jsonify({'ok': False, 'erro': 'Login deve ter pelo menos 3 caracteres.'})
    if len(password) < 4:
        return jsonify({'ok': False, 'erro': 'Senha deve ter pelo menos 4 caracteres.'})

    ok, msg = create_user_file(login, password, nome)
    return jsonify({
        'ok': ok,
        'mensagem': f'Usuário criado. Planilha: usuarios/{login}.xlsx' if ok else msg,
        'erro': None if ok else msg
    })

@app.route('/api/login', methods=['POST'])
def api_login():
    data = request.get_json(silent=True) or {}
    login = safe_login(data.get('login'))
    password = str(data.get('password', '')).strip()

    if login == ADMIN_USER and password == ADMIN_PASS:
        return jsonify({'ok': True, 'tipo': 'admin', 'mensagem': 'Administrador conectado.'})

    ok, msg = validate_user(login, password)
    return jsonify({'ok': ok, 'tipo': 'usuario' if ok else None, 'mensagem': msg, 'erro': None if ok else msg})

@app.route('/api/apostar', methods=['POST'])
def api_apostar():
    data = request.get_json(silent=True) or {}
    login = safe_login(data.get('login'))
    password = str(data.get('password', '')).strip()

    ok, msg = validate_user(login, password)
    if not ok:
        return jsonify({'ok': False, 'erro': msg})

    ok, msg = update_bet(login, data.get('match_id'), data.get('bet_home'), data.get('bet_away'))
    return jsonify({'ok': ok, 'mensagem': msg, 'erro': None if ok else msg})

@app.route('/api/admin/salvar-resultados', methods=['POST'])
def api_admin_salvar_resultados():
    data = request.get_json(silent=True) or {}
    if data.get('admin_user') != ADMIN_USER or data.get('admin_pass') != ADMIN_PASS:
        return jsonify({'ok': False, 'erro': 'Administrador inválido.'})

    save_matches(data.get('matches', []))
    return jsonify({'ok': True, 'mensagem': 'Resultados salvos em dados/jogos.json'})

@app.errorhandler(404)
def not_found(_):
    return jsonify({'ok': False, 'erro': 'Endpoint não encontrado.'}), 404

@app.errorhandler(Exception)
def internal_error(e):
    return jsonify({'ok': False, 'erro': str(e)}), 500

if __name__ == '__main__':
    ensure_dirs()
    load_matches()

    print('===========================================')
    print(' Bolão Copa 2026 Pluma')
    print('===========================================')
    print(f'Rodando localmente em: http://localhost:{PORT}')
    print(f'Rodando na rede em: http://0.0.0.0:{PORT}')
    print('===========================================')

    app.run(host='0.0.0.0', port=PORT)
